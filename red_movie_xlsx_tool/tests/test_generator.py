from __future__ import annotations

from collections import Counter

from openpyxl import load_workbook

from red_movie_xlsx_tool.generator import (
    DEFAULT_QUESTIONS,
    MULTI_SEPARATOR,
    column_title,
    generate_rows,
    generate_workbook,
    is_placeholder_text,
    split_multi,
)

MULTIPLE_IDS = {5, 6, 9, 10, 11, 13, 14, 15, 17}
SINGLE_IDS = {1, 2, 3, 4, 7, 8, 12, 16}


def test_default_schema_has_18_questions_and_correct_types() -> None:
    assert len(DEFAULT_QUESTIONS) == 18
    by_id = {question["id"]: question for question in DEFAULT_QUESTIONS}

    assert {question["id"] for question in DEFAULT_QUESTIONS if question["type"] == "single"} == SINGLE_IDS
    assert {question["id"] for question in DEFAULT_QUESTIONS if question["type"] == "multiple"} == MULTIPLE_IDS
    assert by_id[18]["type"] == "text"

    assert "无聊偶然观看" in by_id[6]["options"]
    assert "面对危险时的无畏勇气与牺牲精神" in by_id[9]["options"]
    assert "情感表达真挚动人" in by_id[10]["options"]
    assert "作用很大，让我对红色精神有了具象化的理解，增强了认同感" in by_id[12]["options"]
    assert "宣传推广力度不足，很多人不知道" in by_id[11]["options"]
    assert by_id[18]["title"] == "您对红色电影助力新时代大学生红色基因传承，还有哪些具体建议？"


def test_multi_separator_is_wjx_split_250b() -> None:
    assert MULTI_SEPARATOR == "┋"
    assert ord(MULTI_SEPARATOR) == 9483


def test_generate_workbook_row_count_and_valid_answers(tmp_path) -> None:
    result = generate_workbook(DEFAULT_QUESTIONS, count=50, seed="demo", output_dir=tmp_path)
    workbook = load_workbook(result.path)
    sheet = workbook["样本答案"]

    assert result.path.name.startswith("red_movie_wjx_split_250b_")
    assert sheet.max_row == 51
    assert sheet.max_column == 19
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:S51"
    assert result.summary["multi_separator"] == "┋"
    assert result.summary["multi_separator_codepoint"] == "U+250B"

    headers = [cell.value for cell in sheet[1]]
    assert headers[0] == "样本编号"

    question_by_header = {column_title(question): question for question in DEFAULT_QUESTIONS}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))
        for header, question in question_by_header.items():
            answer = values[header]
            assert answer
            if question["type"] == "single":
                assert answer in question["options"]
            elif question["type"] == "multiple":
                assert "|" not in answer
                assert ";" not in answer
                assert "；" not in answer
                selected = split_multi(answer)
                assert int(question["min_select"]) <= len(selected) <= int(question["max_select"])
                assert all(item in question["options"] for item in selected)
            else:
                assert not is_placeholder_text(answer)


def test_all_multiple_answers_use_only_wjx_split_250b() -> None:
    rows = generate_rows(DEFAULT_QUESTIONS, count=120, seed="split-250b-check")
    multiple_questions = [question for question in DEFAULT_QUESTIONS if question["type"] == "multiple"]

    joined_values = []
    for row in rows:
        for question in multiple_questions:
            value = row[column_title(question)]
            joined_values.append(value)
            assert "|" not in value
            assert ";" not in value
            assert "；" not in value
            selected = value.split(MULTI_SEPARATOR)
            assert len(selected) == len(split_multi(value))
            assert all(item in question["options"] for item in selected)

    assert any(MULTI_SEPARATOR in value for value in joined_values)


def test_question_11_uses_wjx_split_250b_and_legal_options() -> None:
    rows = generate_rows(DEFAULT_QUESTIONS, count=160, seed="q11-check")
    question = next(item for item in DEFAULT_QUESTIONS if item["id"] == 11)
    header = column_title(question)

    assert any(MULTI_SEPARATOR in row[header] for row in rows)
    for row in rows:
        value = row[header]
        assert "|" not in value
        assert ";" not in value
        assert "；" not in value
        selected = split_multi(value)
        assert int(question["min_select"]) <= len(selected) <= int(question["max_select"])
        assert all(item in question["options"] for item in selected)


def test_generated_rows_are_not_all_duplicates() -> None:
    rows = generate_rows(DEFAULT_QUESTIONS, count=80, seed="variety")
    signatures = {tuple(value for key, value in row.items() if key != "样本编号") for row in rows}
    assert len(signatures) >= 25


def test_seed_is_reproducible() -> None:
    first = generate_rows(DEFAULT_QUESTIONS, count=10, seed="same-seed")
    second = generate_rows(DEFAULT_QUESTIONS, count=10, seed="same-seed")
    assert first == second


def test_no_obvious_placeholder_text() -> None:
    rows = generate_rows(DEFAULT_QUESTIONS, count=60, seed="text-check")
    text_header = column_title(DEFAULT_QUESTIONS[-1])
    for row in rows:
        assert not is_placeholder_text(row[text_header])


def test_distribution_is_reasonable_and_correlated() -> None:
    rows = generate_rows(DEFAULT_QUESTIONS, count=500, seed="distribution-check")
    q1 = column_title(DEFAULT_QUESTIONS[0])
    q2 = column_title(DEFAULT_QUESTIONS[1])
    q3 = column_title(DEFAULT_QUESTIONS[2])
    q4 = column_title(DEFAULT_QUESTIONS[3])
    q8 = column_title(DEFAULT_QUESTIONS[7])
    q13 = column_title(DEFAULT_QUESTIONS[12])

    gender = ratio_counter(rows, q1)
    grade = ratio_counter(rows, q2)
    major = ratio_counter(rows, q3)
    watched = ratio_counter(rows, q4)

    assert 0.38 <= gender["男"] <= 0.54
    assert 0.22 <= grade["大一"] <= 0.34
    assert 0.24 <= grade["大二"] <= 0.39
    assert 0.22 <= major["理工类"] <= 0.38
    assert 0.27 <= watched["完整观看过"] <= 0.41
    assert 0.28 <= watched["看过部分片段"] <= 0.43

    complete_rows = [row for row in rows if row[q4] == "完整观看过"]
    low_rows = [row for row in rows if row[q4] in {"仅听说过未观看", "完全不了解"}]
    positive_answers = {"有非常深刻的理解，深受触动", "有一定理解，对精神内涵有了新认识"}
    complete_positive = sum(1 for row in complete_rows if row[q8] in positive_answers) / len(complete_rows)
    low_positive = sum(1 for row in low_rows if row[q8] in positive_answers) / len(low_rows)
    assert complete_positive > low_positive

    for row in rows:
        selected = split_multi(row[q13])
        if "没有明显影响" in selected:
            assert selected == ["没有明显影响"]


def test_invalid_count_rejected(tmp_path) -> None:
    try:
        generate_workbook(DEFAULT_QUESTIONS, count=0, seed=None, output_dir=tmp_path)
    except ValueError as exc:
        assert "样本数量" in str(exc)
    else:
        raise AssertionError("expected ValueError")


def test_invalid_question_rejected(tmp_path) -> None:
    bad_questions = [{"id": 1, "type": "essay", "title": "bad", "options": ["A"]}]
    try:
        generate_workbook(bad_questions, count=1, seed=None, output_dir=tmp_path)
    except ValueError as exc:
        assert "single、multiple 或 text" in str(exc)
    else:
        raise AssertionError("expected ValueError")


def ratio_counter(rows: list[dict[str, str]], header: str) -> dict[str, float]:
    counts = Counter(row[header] for row in rows)
    return {key: value / len(rows) for key, value in counts.items()}
