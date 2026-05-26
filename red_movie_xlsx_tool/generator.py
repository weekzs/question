from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import random
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SURVEY_TITLE = "红色电影《阮啸仙》对新时代大学生红色基因传承的实践调研问卷"
MULTI_SEPARATOR = "\u250b"
FORBIDDEN_MULTI_SEPARATORS = ("|", ";", "；")
TEXT_PLACEHOLDER_WORDS = ("测试", "占位", "随便", "无", "暂无", "没有")

DEFAULT_QUESTIONS: list[dict[str, Any]] = [
    {
        "id": 1,
        "type": "single",
        "title": "您的性别",
        "options": ["男", "女"],
        "weights": [0.46, 0.54],
    },
    {
        "id": 2,
        "type": "single",
        "title": "您的年级",
        "options": ["大一", "大二", "大三", "大四及以上"],
        "weights": [0.28, 0.32, 0.25, 0.15],
    },
    {
        "id": 3,
        "type": "single",
        "title": "您的专业类别",
        "options": ["理工类", "文史类", "艺术类", "经管类", "其他"],
        "weights": [0.30, 0.26, 0.14, 0.23, 0.07],
    },
    {
        "id": 4,
        "type": "single",
        "title": "您是否观看过红色电影《阮啸仙》？",
        "options": ["完整观看过", "看过部分片段", "仅听说过未观看", "完全不了解"],
        "weights": [0.34, 0.35, 0.24, 0.07],
    },
    {
        "id": 5,
        "type": "multiple",
        "title": "您了解或观看《阮啸仙》的主要渠道是",
        "options": ["学校组织的集体观影活动", "思政课老师推荐", "短视频平台/社交媒体", "家人/朋友推荐", "自行在视频平台观看", "其他"],
        "min_select": 2,
        "max_select": 3,
    },
    {
        "id": 6,
        "type": "multiple",
        "title": "您观看《阮啸仙》的主要原因是",
        "options": ["学校/班级要求参与", "对阮啸仙烈士的事迹感兴趣", "想了解革命历史故事", "想接受红色文化教育", "无聊偶然观看", "其他"],
        "min_select": 2,
        "max_select": 3,
    },
    {
        "id": 7,
        "type": "single",
        "title": "您对影片中阮啸仙烈士的革命事迹了解程度",
        "options": ["非常了解，能说出关键经历", "基本了解，知道大致背景", "了解很少，只知道是革命烈士", "完全不了解"],
    },
    {
        "id": 8,
        "type": "single",
        "title": "观看影片后，您对阮啸仙烈士的革命精神（如坚定信仰、无私奉献、不畏牺牲等）的理解变化",
        "options": ["有非常深刻的理解，深受触动", "有一定理解，对精神内涵有了新认识", "理解变化不大，和之前差不多", "没有明显感受"],
    },
    {
        "id": 9,
        "type": "multiple",
        "title": "您认为影片中最打动您的情节/精神是",
        "options": ["阮啸仙投身革命、坚定信仰的初心", "面对危险时的无畏勇气与牺牲精神", "对群众的关怀与为民情怀", "革命战友间的团结与坚守", "影片的历史还原与细节刻画", "其他"],
        "min_select": 2,
        "max_select": 3,
    },
    {
        "id": 10,
        "type": "multiple",
        "title": "您认为《阮啸仙》作为红色电影，在以下哪些方面做得较好",
        "options": ["历史事实还原度高", "人物形象塑造鲜活立体", "情感表达真挚动人", "叙事节奏紧凑、不枯燥", "对红色精神的传递清晰到位", "其他"],
        "min_select": 2,
        "max_select": 3,
    },
    {
        "id": 11,
        "type": "multiple",
        "title": "您认为影片在哪些方面还有改进空间",
        "options": ["叙事节奏偏慢，缺乏吸引力", "人物刻画不够深入", "历史细节还原不足", "对当代大学生的共鸣点挖掘不够", "宣传推广力度不足，很多人不知道", "其他"],
        "min_select": 1,
        "max_select": 3,
    },
    {
        "id": 12,
        "type": "single",
        "title": "您认为《阮啸仙》这类红色电影，对您了解红色文化、传承红色基因的作用",
        "options": [
            "作用很大，让我对红色精神有了具象化的理解，增强了认同感",
            "有一定作用，让我更愿意主动了解革命历史和红色文化",
            "作用一般，只是看过就忘，没有深入影响",
            "几乎没有作用，对我没什么触动",
        ],
    },
    {
        "id": 13,
        "type": "multiple",
        "title": "观看影片后，您在哪些方面受到了积极影响",
        "options": ["增强了爱国情怀和民族自豪感", "更理解“坚定信仰、勇于担当”的意义", "对革命先辈的奉献精神有了更深的敬意", "更愿意主动学习红色文化知识", "激励自己在学习/生活中更有责任感和奋斗精神", "没有明显影响"],
        "min_select": 1,
        "max_select": 3,
    },
    {
        "id": 14,
        "type": "multiple",
        "title": "您认为新时代大学生传承红色基因，需要做到哪些方面",
        "options": ["主动学习革命历史和红色故事", "践行爱国、敬业、奉献的价值观", "参与红色主题实践活动（如志愿服务、参观纪念馆等）", "向身边人传播红色文化，讲好革命故事", "把红色精神融入日常学习和生活中", "其他"],
        "min_select": 2,
        "max_select": 3,
    },
    {
        "id": 15,
        "type": "multiple",
        "title": "您认为当前高校通过红色电影开展红色文化教育，存在哪些不足",
        "options": ["形式单一，多为集体观影，缺乏互动和讨论", "影片选择有限，部分内容与当代大学生生活脱节", "观影后缺乏延伸学习和引导，难以形成长效影响", "宣传不到位，很多学生不了解这类影片", "其他"],
        "min_select": 1,
        "max_select": 3,
    },
    {
        "id": 16,
        "type": "single",
        "title": "您是否愿意向身边的同学/朋友推荐《阮啸仙》这类红色电影？",
        "options": ["非常愿意，会主动分享并推荐", "愿意，但不会特意推荐", "无所谓，看情况", "不愿意"],
    },
    {
        "id": 17,
        "type": "multiple",
        "title": "您认为可以通过哪些方式，让《阮啸仙》这类红色电影更好地走进大学生群体",
        "options": ["学校多组织观影分享会、主题讨论会", "制作短视频、剧情解说等轻量化内容，在社交平台传播", "推出影片相关的文创产品、互动H5，增强参与感", "结合思政课作业，引导学生撰写观后感或开展调研", "邀请主创或相关学者进校园分享影片背后的故事", "其他"],
        "min_select": 2,
        "max_select": 3,
    },
    {
        "id": 18,
        "type": "text",
        "title": "您对红色电影助力新时代大学生红色基因传承，还有哪些具体建议？",
    },
]


@dataclass(frozen=True)
class GeneratedWorkbook:
    path: Path
    row_count: int
    question_count: int
    summary: dict[str, Any]


def validate_questions(raw_questions: Any) -> list[dict[str, Any]]:
    if not isinstance(raw_questions, list) or not raw_questions:
        raise ValueError("题目 JSON 必须是非空数组")

    questions: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    for index, raw in enumerate(raw_questions, start=1):
        if not isinstance(raw, dict):
            raise ValueError(f"第 {index} 个题目必须是对象")

        question_id = int(raw.get("id", index))
        if question_id in seen_ids:
            raise ValueError(f"题目 ID 重复：{question_id}")
        seen_ids.add(question_id)

        question_type = str(raw.get("type", "")).strip().lower()
        if question_type not in {"single", "multiple", "text"}:
            raise ValueError(f"第 {question_id} 题 type 只能是 single、multiple 或 text")

        title = str(raw.get("title", "")).strip()
        if not title:
            raise ValueError(f"第 {question_id} 题缺少标题")

        question: dict[str, Any] = {"id": question_id, "type": question_type, "title": title}

        if question_type in {"single", "multiple"}:
            options = [str(option).strip() for option in raw.get("options", []) if str(option).strip()]
            if not options:
                raise ValueError(f"第 {question_id} 题至少需要 1 个选项")
            question["options"] = options

            weights = raw.get("weights")
            if weights is not None:
                if not isinstance(weights, list) or len(weights) != len(options):
                    raise ValueError(f"第 {question_id} 题 weights 数量必须和 options 一致")
                question["weights"] = [float(weight) for weight in weights]

        if question_type == "multiple":
            options = question["options"]
            min_select = int(raw.get("min_select", 2))
            max_select = int(raw.get("max_select", min(3, len(options))))
            if min_select < 1 or max_select < min_select or max_select > len(options):
                raise ValueError(f"第 {question_id} 题多选数量范围不合法")
            question["min_select"] = min_select
            question["max_select"] = max_select

        questions.append(question)

    return questions


def generate_rows(raw_questions: Any, count: int, seed: str | int | None = None) -> list[dict[str, str]]:
    if count < 1 or count > 5000:
        raise ValueError("样本数量必须在 1 到 5000 之间")

    questions = validate_questions(raw_questions)
    rng = random.Random(str(seed)) if seed not in (None, "") else random.Random()
    rows: list[dict[str, str]] = []

    for index in range(1, count + 1):
        context: dict[str, str] = {}
        row: dict[str, str] = {"样本编号": f"S{index:04d}"}
        for question in questions:
            answer = answer_question(question, context, rng)
            row[column_title(question)] = answer
            context[str(question["id"])] = answer
        rows.append(row)

    validate_generated_rows(rows, questions)
    return rows


def generate_workbook(
    raw_questions: Any,
    count: int,
    seed: str | int | None,
    output_dir: str | Path,
) -> GeneratedWorkbook:
    questions = validate_questions(raw_questions)
    rows = generate_rows(questions, count, seed)

    output_path = next_output_path(Path(output_dir))
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "样本答案"

    headers = ["样本编号"] + [column_title(question) for question in questions]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    style_sheet(sheet, headers, len(rows))
    workbook.save(output_path)

    return GeneratedWorkbook(
        path=output_path,
        row_count=len(rows),
        question_count=len(questions),
        summary=summarize_rows(rows, headers, questions),
    )


def answer_question(question: dict[str, Any], context: dict[str, str], rng: random.Random) -> str:
    question_type = question["type"]
    if question_type == "single":
        weights = single_weights(question, context)
        return weighted_choice(question["options"], weights, rng)
    if question_type == "multiple":
        return MULTI_SEPARATOR.join(multiple_answers(question, context, rng))
    return text_answer(context, rng)


def single_weights(question: dict[str, Any], context: dict[str, str]) -> list[float]:
    question_id = int(question["id"])
    options = list(question["options"])
    configured = question.get("weights")
    if configured:
        return [float(value) for value in configured]

    viewed = context.get("4", "")
    understanding = context.get("8", "")
    effect = context.get("12", "")

    if question_id == 7:
        if viewed == "完整观看过":
            return weights_for(options, [0.30, 0.50, 0.17, 0.03])
        if viewed == "看过部分片段":
            return weights_for(options, [0.08, 0.45, 0.38, 0.09])
        if viewed == "仅听说过未观看":
            return weights_for(options, [0.02, 0.18, 0.57, 0.23])
        return weights_for(options, [0.01, 0.06, 0.29, 0.64])

    if question_id == 8:
        if viewed == "完整观看过":
            return weights_for(options, [0.38, 0.45, 0.13, 0.04])
        if viewed == "看过部分片段":
            return weights_for(options, [0.15, 0.50, 0.27, 0.08])
        if viewed == "仅听说过未观看":
            return weights_for(options, [0.04, 0.24, 0.46, 0.26])
        return weights_for(options, [0.02, 0.10, 0.31, 0.57])

    if question_id == 12:
        if "非常深刻" in understanding:
            return weights_for(options, [0.64, 0.29, 0.06, 0.01])
        if "一定理解" in understanding:
            return weights_for(options, [0.24, 0.58, 0.15, 0.03])
        if "变化不大" in understanding:
            return weights_for(options, [0.08, 0.28, 0.52, 0.12])
        return weights_for(options, [0.03, 0.12, 0.32, 0.53])

    if question_id == 16:
        if "作用很大" in effect:
            return weights_for(options, [0.52, 0.39, 0.08, 0.01])
        if "有一定作用" in effect:
            return weights_for(options, [0.24, 0.56, 0.18, 0.02])
        if "作用一般" in effect:
            return weights_for(options, [0.06, 0.28, 0.55, 0.11])
        return weights_for(options, [0.03, 0.12, 0.52, 0.33])

    return [1.0 for _ in options]


def multiple_answers(question: dict[str, Any], context: dict[str, str], rng: random.Random) -> list[str]:
    options = list(question["options"])
    question_id = int(question["id"])
    min_select = int(question.get("min_select", 2))
    max_select = int(question.get("max_select", min(3, len(options))))

    if question_id == 13 and low_impact_context(context) and "没有明显影响" in options:
        if rng.random() < 0.58:
            return ["没有明显影响"]

    target_count = choose_multi_count(question_id, min_select, max_select, rng)
    base_weights = multiple_weights(question, context)
    selected: list[str] = []
    available = options[:]
    weights = base_weights[:]

    while available and len(selected) < target_count:
        choice = weighted_choice(available, weights, rng)
        selected.append(choice)
        index = available.index(choice)
        available.pop(index)
        weights.pop(index)

    selected = reduce_other_like(selected, options, base_weights, target_count, rng)

    if question_id == 13 and "没有明显影响" in selected and len(selected) > 1:
        selected = [item for item in selected if item != "没有明显影响"] or ["没有明显影响"]

    return selected[:max_select]


def choose_multi_count(question_id: int, min_select: int, max_select: int, rng: random.Random) -> int:
    if min_select == max_select:
        return min_select

    choices = list(range(min_select, max_select + 1))
    if question_id in {11, 15}:
        preferred = {1: 0.30, 2: 0.48, 3: 0.22}
    else:
        preferred = {1: 0.10, 2: 0.55, 3: 0.35, 4: 0.10}
    weights = [preferred.get(choice, 0.20) for choice in choices]
    return int(weighted_choice([str(choice) for choice in choices], weights, rng))


def multiple_weights(question: dict[str, Any], context: dict[str, str]) -> list[float]:
    options = list(question["options"])
    question_id = int(question["id"])
    viewed = context.get("4", "")
    effect = context.get("12", "")

    if question_id == 5:
        if viewed in {"完整观看过", "看过部分片段"}:
            preferred = {"学校组织的集体观影活动": 1.45, "思政课老师推荐": 1.28, "短视频平台/社交媒体": 1.10, "自行在视频平台观看": 1.02}
        else:
            preferred = {"短视频平台/社交媒体": 1.48, "家人/朋友推荐": 1.18, "思政课老师推荐": 1.04}
        return option_weights(options, preferred, other_weight=0.08)

    if question_id == 6:
        if viewed in {"完整观看过", "看过部分片段"}:
            preferred = {"学校/班级要求参与": 1.25, "对阮啸仙烈士的事迹感兴趣": 1.12, "想了解革命历史故事": 1.28, "想接受红色文化教育": 1.20, "无聊偶然观看": 0.22}
        else:
            preferred = {"学校/班级要求参与": 0.88, "想了解革命历史故事": 0.78, "无聊偶然观看": 1.05}
        return option_weights(options, preferred, other_weight=0.08)

    if question_id == 9:
        preferred = {"阮啸仙投身革命、坚定信仰的初心": 1.25, "面对危险时的无畏勇气与牺牲精神": 1.18, "对群众的关怀与为民情怀": 1.05}
        return option_weights(options, preferred, other_weight=0.07)

    if question_id == 10:
        preferred = {"历史事实还原度高": 1.18, "人物形象塑造鲜活立体": 1.05, "情感表达真挚动人": 1.20, "对红色精神的传递清晰到位": 1.12}
        return option_weights(options, preferred, other_weight=0.07)

    if question_id == 11:
        preferred = {"宣传推广力度不足，很多人不知道": 1.35, "对当代大学生的共鸣点挖掘不够": 1.18, "历史细节还原不足": 1.03, "人物刻画不够深入": 0.92}
        return option_weights(options, preferred, other_weight=0.07)

    if question_id == 13:
        if low_impact_context(context):
            preferred = {"没有明显影响": 1.10, "更愿意主动学习红色文化知识": 0.42}
        else:
            preferred = {
                "增强了爱国情怀和民族自豪感": 1.22,
                "更理解“坚定信仰、勇于担当”的意义": 1.25,
                "对革命先辈的奉献精神有了更深的敬意": 1.15,
                "更愿意主动学习红色文化知识": 0.95,
                "激励自己在学习/生活中更有责任感和奋斗精神": 1.08,
                "没有明显影响": 0.05,
            }
        return option_weights(options, preferred, other_weight=0.07)

    if question_id == 14:
        preferred = {"主动学习革命历史和红色故事": 1.18, "践行爱国、敬业、奉献的价值观": 1.12, "向身边人传播红色文化，讲好革命故事": 1.08}
        return option_weights(options, preferred, other_weight=0.07)

    if question_id == 15:
        if "作用一般" in effect or "几乎没有作用" in effect:
            preferred = {"形式单一，多为集体观影，缺乏互动和讨论": 1.35, "观影后缺乏延伸学习和引导，难以形成长效影响": 1.22, "影片选择有限，部分内容与当代大学生生活脱节": 1.12}
        else:
            preferred = {"宣传不到位，很多学生不了解这类影片": 1.20, "观影后缺乏延伸学习和引导，难以形成长效影响": 1.18, "形式单一，多为集体观影，缺乏互动和讨论": 1.12}
        return option_weights(options, preferred, other_weight=0.07)

    if question_id == 17:
        preferred = {"学校多组织观影分享会、主题讨论会": 1.25, "制作短视频、剧情解说等轻量化内容，在社交平台传播": 1.18, "结合思政课作业，引导学生撰写观后感或开展调研": 1.14, "邀请主创或相关学者进校园分享影片背后的故事": 1.10}
        return option_weights(options, preferred, other_weight=0.07)

    return option_weights(options, {}, other_weight=0.07)


def text_answer(context: dict[str, str], rng: random.Random) -> str:
    effect = context.get("12", "")
    recommend = context.get("16", "")

    practical_topics = [
        "观影后安排班级讨论或心得交流，让同学能把影片内容和自身学习生活联系起来",
        "把红色电影与思政课、主题团日和校园文化活动结合起来，避免只停留在一次性观影",
        "用短视频、海报、推文等新媒体方式提前介绍人物背景，提高同学主动观看的兴趣",
        "结合红色研学、纪念馆参观和社会实践，让影片中的精神有更具体的体验场景",
        "邀请老师、专家或主创人员开展导读分享，帮助同学理解历史背景和人物精神",
        "设置观后感展示、微宣讲或小组汇报，让红色电影教育更有参与感和互动性",
    ]
    low_interest_topics = [
        "建议在观影前先用简短材料介绍阮啸仙烈士的生平，这样观看时更容易理解情节",
        "可以减少单纯播放影片的形式，增加问答、讨论和案例分析，提高参与度",
        "宣传方式可以更贴近大学生习惯，例如结合校园公众号、短视频平台和社团活动推广",
    ]

    pool = low_interest_topics if ("一般" in recommend or "不愿意" in recommend or "作用一般" in effect) else practical_topics
    first = rng.choice(pool)
    if rng.random() < 0.45:
        second_pool = [topic for topic in practical_topics if topic != first]
        second = rng.choice(second_pool)
        return f"{first}。同时，{second}。"
    return f"{first}。"


def validate_generated_rows(rows: list[dict[str, str]], questions: list[dict[str, Any]]) -> None:
    for question in questions:
        if question["type"] != "multiple":
            continue
        header = column_title(question)
        options = set(question["options"])
        min_select = int(question["min_select"])
        max_select = int(question["max_select"])
        for row_index, row in enumerate(rows, start=1):
            value = row.get(header, "")
            if any(separator in value for separator in FORBIDDEN_MULTI_SEPARATORS):
                raise ValueError(f"第 {row_index} 行 {header} 包含非官方多选分隔符")
            selected = split_multi(value)
            if not (min_select <= len(selected) <= max_select):
                raise ValueError(f"第 {row_index} 行 {header} 多选数量不合法")
            invalid = [item for item in selected if item not in options]
            if invalid:
                raise ValueError(f"第 {row_index} 行 {header} 含非法选项：{invalid[0]}")


def weighted_choice(options: list[str], weights: list[float], rng: random.Random) -> str:
    if len(options) != len(weights):
        weights = [1.0 for _ in options]
    total = sum(max(weight, 0.0) for weight in weights)
    if total <= 0:
        return rng.choice(options)
    pick = rng.random() * total
    running = 0.0
    for option, weight in zip(options, weights):
        running += max(weight, 0.0)
        if pick <= running:
            return option
    return options[-1]


def weights_for(options: list[str], weights: list[float]) -> list[float]:
    if len(weights) >= len(options):
        return weights[: len(options)]
    return weights + [1.0] * (len(options) - len(weights))


def option_weights(options: list[str], preferred: dict[str, float], other_weight: float) -> list[float]:
    weights: list[float] = []
    for option in options:
        if is_other_like(option):
            weights.append(other_weight)
        else:
            weights.append(preferred.get(option, 1.0))
    return weights


def reduce_other_like(
    selected: list[str],
    options: list[str],
    base_weights: list[float],
    target_count: int,
    rng: random.Random,
) -> list[str]:
    if not selected_has_other(selected):
        return selected
    if len(options) - count_other_like(options) < target_count:
        return selected
    if rng.random() >= 0.88:
        return selected

    selected = [item for item in selected if not is_other_like(item)]
    non_other = [item for item in options if not is_other_like(item) and item not in selected]
    non_other_weights = [base_weights[options.index(item)] for item in non_other]
    while non_other and len(selected) < target_count:
        choice = weighted_choice(non_other, non_other_weights, rng)
        selected.append(choice)
        index = non_other.index(choice)
        non_other.pop(index)
        non_other_weights.pop(index)
    return selected


def low_impact_context(context: dict[str, str]) -> bool:
    viewed = context.get("4", "")
    understanding = context.get("8", "")
    effect = context.get("12", "")
    return (
        viewed in {"仅听说过未观看", "完全不了解"}
        or "没有明显感受" in understanding
        or "作用一般" in effect
        or "几乎没有作用" in effect
    )


def column_title(question: dict[str, Any]) -> str:
    return f"{question['id']}. {question['title']}"


def selected_has_other(selected: list[str]) -> bool:
    return any(is_other_like(item) for item in selected)


def count_other_like(options: list[str]) -> int:
    return sum(1 for option in options if is_other_like(option))


def is_other_like(option: str) -> bool:
    return option.strip().lower() in {"其他", "其它", "other"}


def is_placeholder_text(value: str) -> bool:
    stripped = value.strip()
    if not stripped or len(stripped) < 8:
        return True
    return stripped in TEXT_PLACEHOLDER_WORDS or any(word in stripped for word in ("测试", "占位", "随便"))


def next_output_path(output_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"red_movie_wjx_split_250b_{timestamp}.xlsx"
    path = output_dir / base_name
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = output_dir / f"red_movie_wjx_split_250b_{timestamp}_{index}.xlsx"
        if not candidate.exists():
            return candidate
    raise ValueError("无法生成唯一文件名")


def summarize_rows(rows: list[dict[str, str]], headers: list[str], questions: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {"columns": len(headers), "unique_answer_rows": 0, "multi_separator": MULTI_SEPARATOR, "single": {}, "multiple": {}, "text": {}}

    signatures = {tuple(row.get(header, "") for header in headers[1:]) for row in rows}
    summary: dict[str, Any] = {
        "columns": len(headers),
        "unique_answer_rows": len(signatures),
        "multi_separator": MULTI_SEPARATOR,
        "multi_separator_codepoint": f"U+{ord(MULTI_SEPARATOR):04X}",
        "single": {},
        "multiple": {},
        "text": {},
        "first_sample": rows[0],
    }

    for question in questions:
        header = column_title(question)
        question_type = question["type"]
        if question_type == "single":
            counts = Counter(row.get(header, "") for row in rows)
            summary["single"][header] = {
                option: {
                    "count": counts.get(option, 0),
                    "percent": round(counts.get(option, 0) / len(rows) * 100, 1),
                }
                for option in question["options"]
            }
        elif question_type == "multiple":
            counts: Counter[str] = Counter()
            for row in rows:
                for item in split_multi(row.get(header, "")):
                    counts[item] += 1
            summary["multiple"][header] = {option: counts.get(option, 0) for option in question["options"]}
        elif question_type == "text":
            filled = sum(1 for row in rows if row.get(header, "").strip())
            summary["text"][header] = {"filled": filled}

    return summary


def split_multi(value: str) -> list[str]:
    return [item for item in value.split(MULTI_SEPARATOR) if item]


def style_sheet(sheet: Any, headers: list[str], row_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="B91C1C")
    header_font = Font(color="FFFFFF", bold=True)
    thin_fill = PatternFill("solid", fgColor="FFF7ED")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2, max_row=row_count + 1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = thin_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count + 1}"

    for column_index, header in enumerate(headers, start=1):
        values = [str(header)]
        for row_index in range(2, min(row_count + 2, 102)):
            values.append(str(sheet.cell(row=row_index, column=column_index).value or ""))
        width = min(max(max(display_width(value) for value in values) + 2, 12), 48)
        if column_index == 1:
            width = 12
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.row_dimensions[1].height = 38


def display_width(value: str) -> int:
    ascii_len = len(re.sub(r"[^\x00-\x7F]", "", value))
    non_ascii_len = len(value) - ascii_len
    return ascii_len + non_ascii_len * 2
